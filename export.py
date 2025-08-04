# export.py
import pandas as pd
from db import load_all_lots
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os

def export_to_excel_rus():
    df = load_all_lots()
    df = df.sort_values(['ann_id', 'plan_point_id'])
    # Русские заголовки
    columns = {
        'ann_id': 'Номер объявления',
        'customer': 'Организатор',
        'plan_point_id': 'Номер лота',
        'title': 'Наименование лота',
        'description': 'Краткая характеристика',
        'quantity': 'Кол-во',
        'amount': 'Сумма',
        'unit': 'Ед.',
        'price': 'Цена за ед.',
        'item_type': 'Вид предмета',
        'date_start': 'Дата начала',
        'date_end': 'Дата окончания',
        'method': 'Способ закупки',
        'status': 'Статус'
    }
    # Только нужные столбцы (порядок!)
    out_cols = [
        'ann_id', 'customer', 'plan_point_id', 'title', 'description', 'quantity', 'amount',
        'unit', 'price', 'item_type', 'date_start', 'date_end', 'method', 'status'
    ]
    df_export = df[out_cols].copy()
    df_export = df_export.rename(columns=columns)
    last_ann_id, last_customer = None, None
    for idx, row in df_export.iterrows():
        if row['Номер объявления'] == last_ann_id:
            df_export.at[idx, 'Номер объявления'] = ''
            df_export.at[idx, 'Организатор'] = ''
        else:
            last_ann_id = row['Номер объявления']
            last_customer = row['Организатор']
    
    os.makedirs("exports", exist_ok=True)
    filename = f"exports/zakupki_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    df_export.to_excel(filename, index=False)

    # Форматирование Excel
    wb = load_workbook(filename)
    ws = wb.active
    ws.freeze_panes = "A2"
    font_bold = Font(bold=True)
    alignment = Alignment(wrap_text=True, vertical="top")

    for col_idx, col in enumerate(ws.iter_cols(1, ws.max_column), 1):
        col[0].font = font_bold
        col[0].alignment = alignment
        # Автоширина (но не больше 50)
        max_length = max(len(str(cell.value)) for cell in col[:100] if cell.value) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length, 50)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = alignment
    wb.save(filename)
    return filename
